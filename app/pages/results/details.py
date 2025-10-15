# app/pages/results/details.py
"""
Detail views and export functionality for simulation results.
"""
import streamlit as st
import pandas as pd
import yaml
from pathlib import Path
from datetime import datetime
from app.models import get_decision_global_parameters, get_all_global_parameters


def render_parameter_applicability_summary():
    """Render parameter applicability summary for the run"""
    selected_decisions = st.session_state.decision_params.selected_decisions
    
    if selected_decisions:
        # Calculate overall applicability
        total_applicable = get_decision_global_parameters(selected_decisions)
        all_global_params = get_all_global_parameters()
        total_not_applicable = all_global_params - total_applicable
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📊 Total Parameters", len(all_global_params))
        with col2:
            st.metric("✅ Applicable", len(total_applicable))
        with col3:
            st.metric("❌ Not Applicable", len(total_not_applicable))
        with col4:
            applicability_pct = len(total_applicable) / len(all_global_params) * 100 if all_global_params else 0
            st.metric("📈 Efficiency", f"{applicability_pct:.0f}%")
        
            # Show which parameters were actually used vs unused
        col_used, col_unused = st.columns(2)
        
        with col_used:
            st.markdown("### ✅ Parameters Used in This Simulation")
            if total_applicable:
                for param in sorted(total_applicable):
                    st.markdown(f"  • {param.replace('_', ' ').title()}")
            else:
                st.caption("No parameters were applicable for the selected decisions.")
        
        with col_unused:
            st.markdown("### ❌ Parameters Not Used in This Simulation")
            if total_not_applicable:
                for param in sorted(total_not_applicable):
                    st.markdown(f"  • {param.replace('_', ' ').title()}")
            else:
                st.caption("All parameters were used in this simulation.")
        
        # Show decision-specific breakdown
        st.markdown("### 📊 Parameter Usage by Decision")
        
        try:
            decisions_path = Path(__file__).resolve().parents[3] / "config" / "decisions.yaml"
            with open(decisions_path, 'r') as f:
                decisions_config = yaml.safe_load(f)
            
            for decision in selected_decisions:
                decision_config = decisions_config.get(decision, {})
                decision_params = set(decision_config.get('uses_global_parameters', []))
                not_used = all_global_params - decision_params
                efficiency = len(decision_params) / len(all_global_params) * 100 if all_global_params else 0
                
                with st.container():
                    col_title, col_metrics = st.columns([2, 3])
                    
                    with col_title:
                        st.markdown(f"**{decision.replace('_', ' ').title()}**")
                    
                    with col_metrics:
                        sub_col1, sub_col2, sub_col3 = st.columns(3)
                        with sub_col1:
                            st.metric("Uses", len(decision_params), label_visibility="collapsed")
                        with sub_col2:
                            st.metric("Doesn't Use", len(not_used), label_visibility="collapsed")
                        with sub_col3:
                            st.metric("Efficiency", f"{efficiency:.0f}%", label_visibility="collapsed")
                    
                    st.markdown("---")
        except Exception as e:
            st.error(f"Error loading decision configurations: {e}")
    else:
        st.caption("No decisions were selected for this simulation.")


def render_individual_agent_details(df):
    """Render individual agent details section"""
    # Check if this is dependent variable mode
    is_depvar_mode = len(df.columns) == 1 and 'donation_default' in df.columns
    
    if not is_depvar_mode:
        st.subheader("🔍 Individual Agent Details")
        
        # Agent selection
        agent_id = st.selectbox(
            "Select Agent to Examine",
            options=range(len(df)),
            format_func=lambda x: f"Agent {x+1}"
        )
        
        agent_data = df.iloc[agent_id]
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("** Agent Traits**")
            trait_data = {}
            for col in ['Honesty_Humility', 'Assigned Allowance Level', 'Study Program', 
                       'Group_experiment', 'TWT+Sospeso [=AW2+AX2]{Periods 1+2}']:
                if col in agent_data:
                    trait_data[col] = agent_data[col]
            
            trait_df = pd.DataFrame(list(trait_data.items()), columns=['Trait', 'Value'])
            # Convert all values to strings to avoid PyArrow serialization issues
            trait_df['Value'] = trait_df['Value'].astype(str)
            st.dataframe(trait_df, hide_index=True)
        
        with col2:
            st.markdown("**🎯 Agent Decisions**")
            decision_data = {}
            for col in df.columns:
                if col not in ['Assigned Allowance Level', 'Group_experiment', 'Honesty_Humility', 
                              'Study Program', 'TWT+Sospeso [=AW2+AX2]{Periods 1+2}']:
                    decision_data[col] = agent_data[col]
            
            decision_df = pd.DataFrame(list(decision_data.items()), columns=['Decision', 'Value'])
            if 'donation_default' in decision_data:
                decision_df.loc[decision_df['Decision'] == 'donation_default', 'Value'] = \
                    f"{decision_data['donation_default']:.2%}"
            # Convert all values to strings to avoid PyArrow serialization issues
            decision_df['Value'] = decision_df['Value'].astype(str)
            st.dataframe(decision_df, hide_index=True)
    else:
        st.caption("Individual agent details not available in dependent variable resampling mode (no trait information)")


def render_export_section(df, results_dict=None, using_selected_config=False):
    """Render the export/download section
    
    Args:
        df: Primary DataFrame for display and single CSV export
        results_dict: Full results dictionary (for multi-sheet Excel export)
        using_selected_config: Whether a configuration was selected
    """
    st.subheader("💾 Export Results")
    
    # Detect if this is an individual donation_default run
    is_donation_only_run = (
        hasattr(st.session_state, 'custom_decisions') and 
        st.session_state.custom_decisions == ['donation_default'] and
        hasattr(st.session_state, 'default_decisions') and
        len(st.session_state.default_decisions) == 0
    )
    
    # Filter columns for donation-only runs
    if is_donation_only_run:
        # Keep only donation_default columns (no traits, no other decisions)
        donation_cols = [col for col in df.columns if 'donation_default' in col or 'donation' in col.lower()]
        df = df[donation_cols]
        
        # Also filter results_dict if provided
        if results_dict:
            filtered_results_dict = {}
            for key, config_df in results_dict.items():
                donation_cols_config = [col for col in config_df.columns if 'donation_default' in col or 'donation' in col.lower()]
                filtered_results_dict[key] = config_df[donation_cols_config]
            results_dict = filtered_results_dict
        
        # Show info message for donation-only exports
        st.info("📊 **Donation Default Comparison Mode:** Excel export includes only Agent_Number and donation_default columns for easy comparison across configurations.")
    
    # Show banner if a specific config is selected
    if using_selected_config and results_dict and len(results_dict) == 1:
        config_name = list(results_dict.keys())[0]
        
        col_msg, col_btn = st.columns([3, 1])
        with col_msg:
            st.success(f"🎯 **Exporting selected configuration:** {config_name.replace('_', ' ').title()}")
            st.caption("To export all configurations, clear the selection first.")
        
        with col_btn:
            from app.pages.decision_execution import clear_selected_configuration
            if st.button("🗑️ Clear Selection", key="export_clear_selection"):
                clear_selected_configuration()
                st.rerun()
    
    # Show a quick verification metric from the SAME DataFrame used for export
    # This helps detect any mismatch between charts and the exported numbers
    if 'donation_default' in df.columns:
        try:
            export_mean = pd.to_numeric(df['donation_default'], errors='coerce').mean()
            colm1, colm2, colm3 = st.columns(3)
            with colm1:
                st.metric("Export Mean (donation_default)", f"{export_mean:.2%}")
            with colm2:
                st.caption("Computed from the exact DataFrame below")
            with colm3:
                st.caption(f"Rows: {len(df):,}")
        except Exception:
            pass
    
    # CRITICAL: Show all donation-related columns in the export to help diagnose confusion
    donation_cols = [col for col in df.columns if 'donation' in col.lower()]
    if donation_cols:
        with st.expander("📊 Donation Columns in Export", expanded=True):
            st.markdown("**This CSV contains the following donation-related columns:**")
            for col in donation_cols:
                try:
                    col_mean = pd.to_numeric(df[col], errors='coerce').mean()
                    if pd.notna(col_mean):
                        st.write(f"• **{col}**: mean = {col_mean:.4f} ({col_mean:.1%})")
                    else:
                        st.write(f"• **{col}**: (non-numeric)")
                except Exception:
                    st.write(f"• **{col}**: (error computing mean)")
            
            st.markdown("---")
            st.info("✅ **Use `donation_default` for the final processed donation rate** (shown in charts above)")
            if 'final_donation_rate' in donation_cols:
                st.caption("⚠️ `final_donation_rate` is a separate decision (slider/default value), not the computed rate")
    
    # Determine export mode
    export_all_configs = (
        results_dict is not None and 
        len(results_dict) > 1 and 
        not using_selected_config
    )
    
    export_single_selected = (
        results_dict is not None and
        len(results_dict) == 1 and
        using_selected_config
    )
    
    if export_all_configs:
        st.info(f"📋 **Multi-Configuration Export Available:** {len(results_dict)} configurations will be exported in a single sheet with separate columns for easy comparison")
        with st.expander("📊 View Configurations to be Exported", expanded=False):
            st.caption("All configurations will be in one sheet with columns suffixed by configuration name")
            st.markdown("**Example columns:** `donation_default_Copula_Categorical`, `donation_default_Research_Spec_Continuous`, etc.")
            st.markdown("---")
            for idx, (config_key, config_df) in enumerate(results_dict.items(), 1):
                col_name, col_metrics = st.columns([2, 2])
                with col_name:
                    st.write(f"**{idx}. {config_key.replace('_', ' ').title()}**")
                with col_metrics:
                    if 'donation_default' in config_df.columns:
                        mean_val = pd.to_numeric(config_df['donation_default'], errors='coerce').mean()
                        st.caption(f"Agents: {len(config_df):,} | Mean donation: {mean_val:.2%}")
                    else:
                        st.caption(f"Agents: {len(config_df):,}")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Excel export - multi-sheet if results_dict provided and no config selected
        try:
            from io import BytesIO
            buffer = BytesIO()
            
            # Determine if we should export all configurations
            export_all_configs = (
                results_dict is not None and 
                len(results_dict) > 1 and 
                not using_selected_config
            )
            
            if export_all_configs:
                # Single sheet Excel with all configurations as separate columns
                # This makes comparison easier than separate sheets
                
                # Get trait columns (same across all configs)
                first_config_df = next(iter(results_dict.values()))
                trait_columns = ['Honesty_Humility', 'Assigned Allowance Level', 'Study Program', 
                                'Group_experiment', 'TWT+Sospeso [=AW2+AX2]{Periods 1+2}']
                
                # Start with Agent_Number and traits from first config (skip traits for donation-only runs)
                if is_donation_only_run:
                    # For donation-only runs, start with empty DataFrame
                    combined_df = pd.DataFrame(index=range(len(first_config_df)))
                    combined_df.insert(0, 'Agent_Number', range(1, len(combined_df) + 1))
                else:
                    # For complete simulation, include trait columns
                    combined_df = first_config_df[trait_columns].copy()
                    combined_df.insert(0, 'Agent_Number', range(1, len(combined_df) + 1))
                
                # Track which columns to color green (donation_default variants)
                green_columns = []
                
                # First pass: Add default decision columns ONCE (they're the same across configs)
                # These are non-donation_default decision columns
                # Skip this for donation-only runs
                default_decision_cols = []
                if not is_donation_only_run:
                    decision_cols_first = [col for col in first_config_df.columns if col not in trait_columns]
                    for col in decision_cols_first:
                        if 'donation_default' not in col:
                            default_decision_cols.append(col)
                            combined_df[col] = first_config_df[col].values
                
                # Second pass: Add donation_default columns from each configuration with config suffix
                for config_key, config_df in results_dict.items():
                    if not config_df.empty:
                        # Get only donation_default related columns
                        decision_cols = [col for col in config_df.columns if col not in trait_columns]
                        
                        # Add only donation_default columns with config suffix
                        for col in decision_cols:
                            if 'donation_default' in col:
                                # Create readable config suffix
                                config_suffix = config_key.replace('_', ' ').title().replace(' ', '_')
                                new_col_name = f"{col}_{config_suffix}"
                                combined_df[new_col_name] = config_df[col].values
                                green_columns.append(new_col_name)
                
                # Export as single sheet with formatting
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    combined_df.to_excel(writer, index=False, sheet_name='All Configurations')
                    
                    # Apply green highlighting to donation_default columns
                    from openpyxl.styles import PatternFill
                    worksheet = writer.sheets['All Configurations']
                    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    
                    # Find column indices for green_columns
                    header_row = list(combined_df.columns)
                    for col_name in green_columns:
                        if col_name in header_row:
                            col_idx = header_row.index(col_name) + 1  # Excel is 1-indexed
                            # Color the header cell
                            cell = worksheet.cell(row=1, column=col_idx)
                            cell.fill = green_fill
                            # Color all data cells in this column
                            for row_idx in range(2, len(combined_df) + 2):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = green_fill
                
                # Adjust label based on whether it's all configs or selected config
                if export_single_selected:
                    excel_label = "📊 Download Excel"
                    excel_filename = f"enhanced_simulation_selected_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    excel_help = "Downloads the selected configuration"
                else:
                    excel_label = f"📊 Download Excel (All {len(results_dict)} Configs)"
                    excel_filename = f"enhanced_simulation_all_configs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    excel_help = f"Downloads all {len(results_dict)} configurations in one sheet with separate columns for easy comparison"
                
                st.download_button(
                    label=excel_label,
                    data=buffer.getvalue(),
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help=excel_help
                )
            else:
                # Single sheet Excel
                # Add Agent Number as first column
                df_export = df.copy()
                df_export.insert(0, 'Agent_Number', range(1, len(df_export) + 1))
                
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_export.to_excel(writer, index=False, sheet_name='Results')
                
                st.download_button(
                    label="📊 Download Excel",
                    data=buffer.getvalue(),
                    file_name=f"enhanced_simulation_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except ImportError:
            st.caption("⚠️ Excel export requires openpyxl")
            st.caption("Install with: pip install openpyxl")
    
    with col2:
        # Transaction-level export (if purchase_requests exist)
        if 'purchase_requests' in df.columns:
            try:
                from io import BytesIO
                
                # Flatten purchase_requests to transaction-level DataFrame
                transactions = []
                transaction_id = 1
                
                for idx, row in df.iterrows():
                    purchase_requests = row.get('purchase_requests', [])
                    if isinstance(purchase_requests, list):
                        for req in purchase_requests:
                            if isinstance(req, dict):
                                transactions.append({
                                    'transaction_id': transaction_id,
                                    'customer_id': req.get('customer_id', idx + 1),
                                    'vendorID': req.get('vendorID', 1),
                                    'platformPrice': req.get('platformPrice', 'N/A'),
                                    'purchase_bid_value': req.get('bid_value', 'N/A'),
                                    'timestamp': req.get('timestamp_hours', 0.0)
                                })
                                transaction_id += 1
                
                if len(transactions) > 0:
                    transactions_df = pd.DataFrame(transactions)
                    
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        transactions_df.to_excel(writer, index=False, sheet_name='Transactions')
                    
                    st.download_button(
                        label="📋 Download Transactions",
                        data=buffer.getvalue(),
                        file_name=f"transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Transaction-level data: one row per purchase request"
                    )
                    
                    # Show preview
                    with st.expander("📋 Transaction Preview", expanded=False):
                        st.dataframe(transactions_df.head(50), use_container_width=True)
                        st.caption(f"Showing first 50 of {len(transactions_df):,} total transactions")
                else:
                    st.caption("No transactions to export")
            
            except ImportError:
                st.caption("⚠️ Transaction export requires openpyxl")
    
    # Add clear button in a new row
    if st.button("🔄 Clear Results"):
        st.session_state.simulation_results = None
        st.rerun()
